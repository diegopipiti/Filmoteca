import decimal
from pathlib import Path

import requests
from requests import HTTPError
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from films.models import Director, Film, Genre


def parse_rating(value: str):
    if not value or value == 'N/A':
        return None
    try:
        number = decimal.Decimal(value)
        return number.quantize(decimal.Decimal('0.1'))
    except (decimal.InvalidOperation, ValueError):
        return None


def normalize_list(raw: str):
    if not raw or raw == 'N/A':
        return []
    return [item.strip() for item in raw.split(',') if item.strip()]


def fetch_metadata(title: str, api_key: str):
    response = requests.get(
        'https://www.omdbapi.com/',
        params={'t': title, 'apikey': api_key, 'type': 'movie'},
        timeout=10,
    )
    response.raise_for_status()
    payload = response.json()

    if payload.get('Response') != 'True':
        raise ValueError(payload.get('Error') or 'Titolo non trovato')

    return {
        'year': payload.get('Year'),
        'director': payload.get('Director'),
        'genres': payload.get('Genre'),
        'rating': payload.get('imdbRating'),
        'poster': payload.get('Poster'),
        'plot': payload.get('Plot'),
    }


def infer_title_column(sheet):
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, cell in enumerate(header):
        if not cell:
            continue
        value = str(cell).strip().lower()
        if value in {'titolo', 'title', 'film'}:
            return idx
    return 0


def validate_api_key(api_key: str):
    """Esegue una chiamata di verifica con un titolo noto per intercettare subito chiavi errate/inattive."""

    try:
        response = requests.get(
            'https://www.omdbapi.com/',
            params={'t': 'Matrix', 'apikey': api_key, 'type': 'movie'},
            timeout=10,
        )
        response.raise_for_status()
    except HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 401:
            raise CommandError(
                'OMDb ha risposto 401 Unauthorized: verifica che OMDB_API_KEY in .env sia corretta, attiva e non sia '
                'più il segnaposto. Puoi testarla anche con `curl "https://www.omdbapi.com/?t=Matrix&apikey=$OMDB_API_KEY"`.'
            ) from exc
        raise CommandError(f"Errore HTTP durante la verifica della chiave OMDb: {exc}") from exc
    except Exception as exc:  # broad per timeout / problemi di rete
        raise CommandError(f"Impossibile verificare la chiave OMDb: {exc}") from exc

    payload = response.json()
    if payload.get('Response') != 'True':
        raise CommandError(
            "La chiave OMDb ha risposto ma non sembra attiva/valida (titolo di prova non trovato). "
            "Controlla che la chiave sia stata attivata via email sul sito OMDb."
        )


class Command(BaseCommand):
    help = 'Importa film da un file XLSX che contiene almeno il titolo. Usa OMDb per completare gli altri campi.'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Percorso del file XLSX con i titoli')
        parser.add_argument(
            '--watched', action='store_true', help='Imposta flag "visto" a True per tutti i film importati'
        )
        parser.add_argument(
            '--titles-only',
            action='store_true',
            help='Importa solo i titoli senza completamento OMDb (altri campi vuoti)',
        )

    def handle(self, *args, **options):
        file_path = Path(options['file_path'])
        watched_default = options['watched']
        titles_only = options['titles_only']
        api_key = getattr(settings, 'OMDB_API_KEY', None)

        if not titles_only:
            if not api_key:
                raise CommandError('Configura OMDB_API_KEY nel file .env prima di eseguire questo comando.')

            if api_key.upper() == 'INSERISCI_LA_TUA_CHIAVE':
                raise CommandError('Sostituisci INSERISCI_LA_TUA_CHIAVE con la tua API key OMDb reale in .env.')

            validate_api_key(api_key)

        if not file_path.exists():
            raise CommandError(f"File non trovato: {file_path}")

        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        title_col = infer_title_column(sheet)
        imported = 0
        skipped = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_title = row[title_col]
            if not raw_title:
                skipped.append((idx, 'Titolo mancante'))
                continue

            title = str(raw_title).strip()
            if titles_only:
                film, created = Film.objects.get_or_create(
                    title=title,
                    defaults={'watched': watched_default},
                )
                film.genres.clear()
            else:
                try:
                    metadata = fetch_metadata(title, api_key)
                except HTTPError as exc:
                    if exc.response is not None and exc.response.status_code == 401:
                        raise CommandError(
                            'OMDb ha risposto 401 Unauthorized: verifica che OMDB_API_KEY in .env sia corretta '
                            'e attiva. Puoi testarla con `curl "https://www.omdbapi.com/?t=Matrix&apikey=$OMDB_API_KEY"`.'
                        ) from exc
                    skipped.append((idx, f"Errore API HTTP: {exc}"))
                    continue
                except Exception as exc:  # broad for API/HTTP errors
                    skipped.append((idx, f"Errore API: {exc}"))
                    continue

                year_raw = metadata.get('year')
                try:
                    year = int(str(year_raw)[:4])
                except (TypeError, ValueError):
                    skipped.append((idx, 'Anno non valido'))
                    continue

                directors = normalize_list(metadata.get('director'))
                director_name = directors[0] if directors else 'Sconosciuto'
                director, _ = Director.objects.get_or_create(name=director_name)

                film, created = Film.objects.get_or_create(
                    title=title,
                    year=year,
                    director=director,
                    defaults={
                        'watched': watched_default,
                        'rating': parse_rating(metadata.get('rating')),
                        'poster_url': metadata.get('poster') if metadata.get('poster') not in {None, 'N/A'} else '',
                        'notes': metadata.get('plot') if metadata.get('plot') not in {None, 'N/A'} else '',
                    },
                )

                genre_names = normalize_list(metadata.get('genres'))
                if genre_names:
                    genres = [Genre.objects.get_or_create(name=name)[0] for name in genre_names]
                    film.genres.set(genres)
                else:
                    film.genres.clear()

            if created:
                imported += 1
            else:
                message = "Film già presente"
                if not titles_only:
                    message += ", aggiornati solo i generi"
                self.stdout.write(self.style.NOTICE(f"{message}: {title}"))

        summary = f"Import completato: {imported} inseriti, {len(skipped)} saltati."
        self.stdout.write(self.style.SUCCESS(summary))

        if skipped:
            for row_index, reason in skipped:
                self.stdout.write(self.style.WARNING(f"Riga {row_index}: {reason}"))
